'''
Copyright (c) 2010 openpyxl

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in
all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
THE SOFTWARE.

@license: http://www.opensource.org/licenses/mit-license.php
@author: Eric Gazoni
'''

from tests.helper import BaseTestCase

from openpyxl.workbook import Workbook
from openpyxl.worksheet import Worksheet

class TestWorksheet(BaseTestCase):

    def setUp(self):

        self.wb = Workbook()

    def test_new_worksheet(self):

        ws = Worksheet(parent_workbook = self.wb)

    def test_get_cell(self):

        ws = Worksheet(parent_workbook = self.wb)

        c = ws.cell(coordinate = 'A1')

        self.assertEqual(c.get_coordinate(), 'A1')

    def test_set_wrong_title(self):

        self.assertRaises(Exception, Worksheet, self.wb, 'X' * 50)

    def test_worksheet_dimension(self):

        ws = Worksheet(parent_workbook = self.wb)

        self.assertEqual('A1:A1', ws.calculate_dimension())

        ws.cell('B12').value = 'AAA'

        self.assertEqual('A1:B12', ws.calculate_dimension())

